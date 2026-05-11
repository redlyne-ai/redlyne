"""
Build a permanent Excel record of every Redlyne benchmark result —
both detection and remediation.

Usage:
    python tests/build_benchmark_xlsx.py

Reads the raw numbers in this script (snapshots of the runs on
2026-05-11) and emits benchmarks/excel_archive/benchmark.xlsx with
five sheets:

  1. Detection — Headline      block per dataset, every metric per tool
  2. Detection — Full data     long format, filterable
  3. Remediation — Headline    block per dataset, every metric per tool
  4. Remediation — Full data   long format, filterable
  5. Methodology               notes on what each column means

Update the SNAPSHOT blocks at the top (ROWS for detection, REM_ROWS for
remediation) after a new bench run, then re-run.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT = REPO_ROOT / "benchmarks" / "excel_archive" / "benchmark.xlsx"

# ---------------------------------------------------------------------------
# SNAPSHOT — paste here the numbers from `bench_baselines.py` whenever you
# want to refresh the workbook. Numbers below are from run on 2026-05-11.
#
# Each row carries every metric we report; paired datasets fill in P/R/F1/Acc,
# vuln-only datasets leave precision/F1/accuracy as None.
# ---------------------------------------------------------------------------
ROWS = [
    # (dataset, mode, tool, cov_pct, analyzed, total,
    #  p_all, r_all, f1_all, acc_all,
    #  p_an,  r_an,  f1_an,  acc_an,
    #  ms_per_file)

    # PoisonPy — paired
    ("PoisonPy",     "paired",    "Bandit",    0.17, 53,  310, 0.692, 0.058, 0.107, 0.516, 0.692, 0.300, 0.419, 0.528,  19.6),
    ("PoisonPy",     "paired",    "Semgrep",   0.86, 267, 310, 0.696, 0.206, 0.318, 0.558, 0.696, 0.221, 0.335, 0.524, 699.8),
    ("PoisonPy",     "paired",    "Pylint",    0.17, 53,  310, 0.558, 0.187, 0.280, 0.519, 0.558, 0.967, 0.707, 0.547,  58.7),
    ("PoisonPy",     "paired",    "DeVAIC v2 stock", 1.00, 310, 310, 0.680, 0.645, 0.662, 0.671, 0.680, 0.645, 0.662, 0.671, 0.5),
    ("PoisonPy",     "paired",    "Redlyne v0.1.2",  1.00, 310, 310, 0.714, 0.968, 0.822, 0.790, 0.714, 0.968, 0.822, 0.790, 1.4),

    # SecurityEval — vuln-only
    ("SecurityEval", "vuln_only", "Bandit",    1.00, 121, 121, None, 0.405, None, None, None, 0.405, None, None,  20.4),
    ("SecurityEval", "vuln_only", "Semgrep",   1.00, 121, 121, None, 0.347, None, None, None, 0.347, None, None, 710.7),
    ("SecurityEval", "vuln_only", "Pylint",    1.00, 121, 121, None, 0.595, None, None, None, 0.595, None, None,  69.9),
    ("SecurityEval", "vuln_only", "DeVAIC v2 stock", 1.00, 121, 121, None, 0.636, None, None, None, 0.636, None, None, 1.2),
    ("SecurityEval", "vuln_only", "Redlyne v0.1.2",  1.00, 121, 121, None, 0.934, None, None, None, 0.934, None, None, 2.6),

    # Copilot CWE Scenarios — vuln-only (sampled ≤5 per scenario)
    ("Copilot",      "vuln_only", "Bandit",    1.00, 150, 150, None, 0.847, None, None, None, 0.847, None, None,  19.9),
    ("Copilot",      "vuln_only", "Semgrep",   1.00, 150, 150, None, 0.513, None, None, None, 0.513, None, None, 733.7),
    ("Copilot",      "vuln_only", "Pylint",    1.00, 150, 150, None, 0.933, None, None, None, 0.933, None, None, 108.5),
    ("Copilot",      "vuln_only", "DeVAIC v2 stock", 1.00, 150, 150, None, 0.680, None, None, None, 0.680, None, None, 2.8),
    ("Copilot",      "vuln_only", "Redlyne v0.1.2",  1.00, 150, 150, None, 0.893, None, None, None, 0.893, None, None, 6.7),

    # SafeCoder — paired vuln + commit-fix
    ("SafeCoder",    "paired",    "Bandit",    0.44, 467,  1052, 0.776, 0.302, 0.435, 0.607, 0.776, 0.682, 0.726, 0.743,  20.1),
    ("SafeCoder",    "paired",    "Semgrep",   1.00, 1052, 1052, 0.851, 0.369, 0.515, 0.652, 0.851, 0.369, 0.515, 0.652, 706.5),
    ("SafeCoder",    "paired",    "Pylint",    0.44, 467,  1052, 0.498, 0.409, 0.449, 0.498, 0.498, 0.923, 0.647, 0.497,  65.8),
    ("SafeCoder",    "paired",    "DeVAIC v2 stock", 1.00, 1052, 1052, 0.546, 0.462, 0.501, 0.539, 0.546, 0.462, 0.501, 0.539, 2.3),
    ("SafeCoder",    "paired",    "Redlyne v0.1.2",  1.00, 1052, 1052, 0.550, 0.561, 0.556, 0.551, 0.550, 0.561, 0.556, 0.551, 4.5),

    # PromSec — vuln-only
    ("PromSec",      "vuln_only", "Bandit",    1.00, 600, 600, None, 0.928, None, None, None, 0.928, None, None,  20.5),
    ("PromSec",      "vuln_only", "Semgrep",   1.00, 600, 600, None, 0.870, None, None, None, 0.870, None, None, 686.5),
    ("PromSec",      "vuln_only", "Pylint",    1.00, 600, 600, None, 0.988, None, None, None, 0.988, None, None,  98.6),
    ("PromSec",      "vuln_only", "DeVAIC v2 stock", 1.00, 600, 600, None, 0.852, None, None, None, 0.852, None, None, 4.0),
    ("PromSec",      "vuln_only", "Redlyne v0.1.2",  1.00, 600, 600, None, 0.970, None, None, None, 0.970, None, None, 10.3),

    # CVEfixes — paired CVE-file pairs (PRIVATE — DO NOT publish externally
    # until rule coverage on real-world CVEs is broadened; Redlyne is only
    # +0.01 F1 vs DeVAIC v2 stock on this set).
    ("CVEfixes",     "paired",    "Bandit",    0.914, 2816, 3082, 0.509, 0.382, 0.437, 0.507, 0.509, 0.418, 0.459, 0.507,  21.0),
    ("CVEfixes",     "paired",    "Semgrep",   0.934, 2880, 3082, 0.600, 0.178, 0.275, 0.530, 0.600, 0.191, 0.290, 0.532, 659.1),
    ("CVEfixes",     "paired",    "Pylint",    0.914, 2816, 3082, 0.499, 0.836, 0.625, 0.498, 0.499, 0.913, 0.645, 0.497, 171.7),
    ("CVEfixes",     "paired",    "DeVAIC v2 stock", 1.00,  3082, 3082, 0.509, 0.487, 0.498, 0.509, 0.509, 0.487, 0.498, 0.509,  10.7),
    ("CVEfixes",     "paired",    "Redlyne v0.1.2",  1.00,  3082, 3082, 0.510, 0.504, 0.507, 0.509, 0.510, 0.504, 0.507, 0.509,  20.8),
]

DATASETS = ["PoisonPy", "SecurityEval", "Copilot", "SafeCoder", "PromSec", "CVEfixes"]
TOOLS    = ["Bandit", "Semgrep", "Pylint", "DeVAIC v2 stock", "Redlyne v0.1.2"]
PAIRED   = {"PoisonPy", "SafeCoder", "CVEfixes"}
DATASET_META = {
    "PoisonPy":     ("Cotroneo et al., ICPC 2024",     "Synthetic paired vuln + clean (155+155). Supports P/R/F1/Acc."),
    "SecurityEval": ("Siddiq & Santos, MSR 2022",      "Hand-curated vulnerable Python by CWE (121). Recall only — no clean samples."),
    "Copilot":      ("Pearce et al., S&P 2022",        "Copilot-generated Python in CWE scenarios. ≤5/scenario sampled (150). Recall only."),
    "SafeCoder":    ("He et al., ICML 2024",           "Real commit-based paired before/after (526+526). Supports P/R/F1/Acc."),
    "PromSec":      ("Nazzal et al., CCS 2024",        "Copilot-generated vulnerable Python (500 train + 100 test). Recall only."),
    "CVEfixes":     ("Bhandari et al., PROMISE 2021",  "Real CVE fixes from public OSS, paired code_before/code_after (1541+1541). Supports P/R/F1/Acc."),
}

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
FONT_BODY   = Font(name="Arial", size=11)
FONT_BOLD   = Font(name="Arial", size=11, bold=True)
FONT_HEAD   = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_TITLE  = Font(name="Arial", size=14, bold=True)
FILL_HEAD   = PatternFill("solid", start_color="0F3D5C")
FILL_REDLY  = PatternFill("solid", start_color="E6F4FF")   # highlight Redlyne column
FILL_DEVAIC = PatternFill("solid", start_color="F4F0FB")   # highlight DeVAIC peer column
FILL_NOTE   = PatternFill("solid", start_color="FFF8E1")
THIN        = Side(border_style="thin", color="C0C0C0")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEAD
        cell.fill = FILL_HEAD
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def style_data_cell(cell, fill=None, num_format=None, align="center"):
    cell.font = FONT_BODY
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = BORDER
    if fill is not None:
        cell.fill = fill
    if num_format is not None:
        cell.number_format = num_format


def autosize(ws, min_w=10, max_w=42):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = min_w
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            width = max(width, min(max_w, len(str(v)) + 2))
        ws.column_dimensions[letter].width = width


# ---------------------------------------------------------------------------
# Sheet 1 — Headline: one block per dataset, every metric per tool column
# ---------------------------------------------------------------------------
# Metric rows present in each dataset block. Tuples are (label, tuple-index,
# numeric format, available-for-vuln-only? bool).
#
# Tuple indexes match the ROWS schema:
#   0 dataset  1 mode  2 tool  3 cov  4 analyzed  5 total
#   6 p_all    7 r_all  8 f1_all   9 acc_all
#   10 p_an    11 r_an  12 f1_an   13 acc_an
#   14 ms_per_file
METRIC_ROWS = [
    ("Coverage",              3,  "0%",     True),
    ("Precision (all)",       6,  "0.0%",   False),
    ("Recall (all)",          7,  "0.0%",   True),
    ("F1 (all)",              8,  "0.000",  False),
    ("Accuracy (all)",        9,  "0.0%",   False),
    ("Precision (analyzed)",  10, "0.0%",   False),
    ("Recall (analyzed)",     11, "0.0%",   True),
    ("F1 (analyzed)",         12, "0.000",  False),
    ("Accuracy (analyzed)",   13, "0.0%",   False),
    ("Speed (ms / file)",     14, "0.0",    True),
]


def build_headline(wb: Workbook):
    ws = wb.create_sheet("Detection — Headline")
    ws["A1"] = "Detection — every metric per dataset × tool"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TOOLS) + 1)
    ws["A2"] = ("One block per dataset. 'all' = operational metrics (parse failures count as miss). "
                "'analyzed' = metrics restricted to samples the tool actually parsed. "
                "Vulnerable-only datasets (SecurityEval, Copilot, PromSec) only have Recall measurable.")
    ws["A2"].font = FONT_BODY
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(TOOLS) + 1)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    lookup = {(r[0], r[2]): r for r in ROWS}

    cur_row = 4
    for ds in DATASETS:
        is_paired = ds in PAIRED
        cite, descr = DATASET_META[ds]

        # Dataset header
        title_cell = ws.cell(row=cur_row, column=1,
                             value=f"{ds}  —  {cite}")
        title_cell.font = FONT_HEAD
        title_cell.fill = FILL_HEAD
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=len(TOOLS) + 1)
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Dataset description
        desc_cell = ws.cell(row=cur_row, column=1, value=descr)
        desc_cell.font = FONT_BODY
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=len(TOOLS) + 1)
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

        # Tool header row for this block
        ws.cell(row=cur_row, column=1, value="Metric")
        for c, tool in enumerate(TOOLS, start=2):
            ws.cell(row=cur_row, column=c, value=tool)
        style_header_row(ws, cur_row, len(TOOLS) + 1)
        cur_row += 1

        # Metric rows
        for metric_label, tup_idx, fmt, available_vuln_only in METRIC_ROWS:
            # Skip P/F1/Acc on vuln-only datasets — they're not measurable
            if not is_paired and not available_vuln_only:
                continue
            ws.cell(row=cur_row, column=1, value=metric_label)
            ws.cell(row=cur_row, column=1).font = FONT_BOLD
            ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=cur_row, column=1).border = BORDER
            for c, tool in enumerate(TOOLS, start=2):
                r = lookup.get((ds, tool))
                v = r[tup_idx] if r else None
                cell = ws.cell(row=cur_row, column=c, value=v)
                fill = FILL_REDLY if tool.startswith("Redlyne") else (
                    FILL_DEVAIC if tool.startswith("DeVAIC") else None
                )
                style_data_cell(cell, fill=fill, num_format=fmt)
            cur_row += 1

        # Spacer row between dataset blocks
        cur_row += 1

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36
    # Column widths: first column wider for metric labels, others uniform
    ws.column_dimensions["A"].width = 24
    for c in range(2, len(TOOLS) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.freeze_panes = "B4"


# ---------------------------------------------------------------------------
# Sheet 2 — Full long-format table with every metric
# ---------------------------------------------------------------------------
def build_full(wb: Workbook):
    ws = wb.create_sheet("Detection — Full data")
    ws["A1"] = "Detection — full long-format results"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)

    headers = [
        "Dataset", "Mode", "Tool",
        "Coverage", "Analyzed", "Total",
        "Precision (all)", "Recall (all)", "F1 (all)", "Accuracy (all)",
        "Precision (analyzed)", "Recall (analyzed)", "F1 (analyzed)", "Accuracy (analyzed)",
        "Speed (ms/file)",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))

    for i, r in enumerate(ROWS, start=4):
        for c, v in enumerate(r, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            tool = r[2]
            fill = FILL_REDLY if tool.startswith("Redlyne") else (
                FILL_DEVAIC if tool.startswith("DeVAIC") else None
            )
            # Number formatting per column
            if c == 4:          # coverage
                fmt = "0%"
                style_data_cell(cell, fill=fill, num_format=fmt)
            elif c in (5, 6):   # analyzed, total
                style_data_cell(cell, fill=fill, num_format="0")
            elif c in (7, 8, 10, 11, 12, 14):   # P, R, Acc as %
                style_data_cell(cell, fill=fill, num_format="0.0%")
            elif c in (9, 13):  # F1 as decimal
                style_data_cell(cell, fill=fill, num_format="0.000")
            elif c == 15:        # ms
                style_data_cell(cell, fill=fill, num_format="0.0")
            elif c in (1, 2, 3): # text
                cell.font = FONT_BOLD if c in (1, 3) else FONT_BODY
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = BORDER
                if fill is not None:
                    cell.fill = fill

    # Add an Excel filter to the header row
    ws.auto_filter.ref = f"A3:O{3 + len(ROWS)}"
    ws.freeze_panes = "D4"
    autosize(ws, max_w=22)


# ---------------------------------------------------------------------------
# Sheet 3 — Methodology
# ---------------------------------------------------------------------------
def build_methodology(wb: Workbook):
    ws = wb.create_sheet("Methodology")
    ws["A1"] = "Methodology notes"
    ws["A1"].font = FONT_TITLE

    lines = [
        "",
        "Datasets",
        "--------",
    ]
    for ds in DATASETS:
        cite, descr = DATASET_META[ds]
        lines.append(f"• {ds} ({cite}) — {descr}")
    lines.extend([
        "",
        "Tools",
        "-----",
        "• Bandit — Python AST-based static analyzer. Operational coverage tanks on syntactically informal AI code.",
        "• Semgrep — auto rules registry, taint mode where applicable.",
        "• Pylint — `errors+fatal` severity only (we exclude lint-style warnings to make precision comparable).",
        "• DeVAIC v2 stock — same engine as Redlyne, original (Cotroneo et al.) rule set. POSIX-compatible subset of the 441 rules (~320 of them).",
        "• Redlyne v0.1.2 — 459 rules: DeVAIC v2 + 14 multi-line template rules + Jinja2/sha256-password/path-join coverage.",
        "",
        "Two-metric reporting",
        "--------------------",
        "Each row reports both 'operational' and 'analyzed-only' metrics:",
        "  • Operational (all samples) — parse failures count as misses. This is what an end user experiences in their editor.",
        "  • On analyzed (subset)      — restricted to the samples the tool was actually able to parse. Matches the metric definition in the original baseline papers; favorable to AST-based tools because their parse-failure cliff is hidden.",
        "When publishing externally we lead with the operational table and disclose the coverage column. The analyzed-only numbers are kept for academic apples-to-apples comparison.",
        "",
        "Caveats — detection",
        "-------------------",
        "• CVEfixes — real-world OSS CVEs. Redlyne v0.1.2 lands at F1=0.507 vs DeVAIC v2 stock at F1=0.498 (+0.009). The narrow gap reflects how many CWE classes in real-world CVEs aren't yet covered by our rule set. Useful as an internal yardstick; not yet at the level we'd publish externally.",
        "• PromSec contains 100 'Testing_DS' files without explicit CWE labels — those count toward coverage but their CWE is recorded as '?'.",
        "• Pylint 'errors+fatal' on SafeCoder shows Acc=49.7% on the analyzed subset — that's a random classifier. The 92.3% recall is offset by the same magnitude in false positives. Same pattern on CVEfixes: 91.3% recall, 49.7% accuracy — Pylint flags almost everything.",
        "• Bandit / Pylint coverage cliff: 17% on PoisonPy, 44% on SafeCoder, ~91% on CVEfixes — their AST-based parsers give up on syntactically informal AI / pre-commit-fix Python. That gap is the operational cost of a non-resilient parser.",
        "",
        "Remediation lineup",
        "------------------",
        "DeVAIC v2 stock is intentionally EXCLUDED from the remediation comparison: it ships only 2 remediation rules out of 441 (0.5%), so it's not a remediation tool. It remains a peer in the detection benchmark, where it's directly comparable on P/R/F1/Acc. The remediation comparison is Redlyne (71/459 rules with remediation = 15%) vs PatchitPy (same lineage, full remediation pipeline) vs Semgrep --autofix (~5% of registry has `fix:` blocks).",
        "",
        "Remediation metrics",
        "-------------------",
        "Five per-patch checks compose two 'fully clean' definitions:",
        "  • Applied         — the tool changed the source (vs declining / silent).",
        "  • Syntax-safe     — patched source still compiles.",
        "  • Targeted-clean  — the specific rule IDs that fired pre-patch AND carry a remediation block stop firing post-patch. The honest 'did the fix work?' metric: taint-source / detection-only rules without remediation are excluded because the engine never promised to fix those.",
        "  • Regression-free — no NEW rule IDs appear in the patched source.",
        "  • Cross-clean     — strictest: post-patch source has zero findings overall. Penalizes single-fix tools when the original file carries multiple unrelated vulnerabilities.",
        "Composite:",
        "  • Targeted-full = Syntax-safe AND Regression-free AND Targeted-clean — HEADLINE metric, 'did the patch fix what it targeted?'",
        "  • Strict-full   = Syntax-safe AND Regression-free AND Cross-clean   — strict stress test.",
        "",
        "Caveats — remediation",
        "---------------------",
        "• PatchitPy applied=0% on both datasets — the bash subprocess pipeline did not produce successful patches in our setup. To investigate before publishing externally; possibly an environment / dependency issue with our wrapper. Reported faithfully for now.",
        "• Semgrep --autofix on SafeCoder: applied=2.9% but targeted-clean of applied=6.7%. The autofix removes one specific finding, but the bench's targeted check measures whether Redlyne's rules — which see *more* — confirm the fix. Most Semgrep autofixes here are cosmetic or partial in Redlyne's view.",
        "• Redlyne SafeCoder is lower than PoisonPy across the board because the dataset contains real production-grade fixes with significant function refactors. Our regex-based remediation handles drop-in substitutions well, not structural rewrites.",
        "• CVEfixes — no remediation bench yet. PoisonPy + SafeCoder are the two datasets with paired ground-truth fixes.",
    ])

    for i, line in enumerate(lines, start=2):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = FONT_BOLD if line and line[0].isalpha() and not line.startswith("•") and not line.startswith("  ") else FONT_BODY
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if line.startswith("•") or line.startswith("  "):
            cell.fill = FILL_NOTE

    ws.column_dimensions["A"].width = 130
    ws.row_dimensions[1].height = 22


# ===========================================================================
# REMEDIATION BENCHMARK
# ===========================================================================
# Tools (PatchitPy + Semgrep + Redlyne; DeVAIC v2 excluded — 2/441 rules
# with remediation = not a remediation tool).
REM_TOOLS = ["Semgrep --autofix", "PatchitPy (bash)", "Redlyne v0.1.2"]
REM_DATASETS = ["PoisonPy", "SafeCoder"]  # CVEfixes not benched for remediation

# Snapshot from run 2026-05-11 13:22 (benchmarks/remediation_results/
# run_20260511_132251.json). Schema:
#   (dataset, tool, total, applied, syntax_safe, targeted_clean,
#    regression_free, cross_clean, targeted_full, strict_full,
#    applied_rate, rate_safe_of_applied,
#    rate_targeted_of_applied, rate_targeted_of_total,
#    rate_clean_of_applied,    rate_clean_of_total,
#    mean_similarity_to_truth, ms_per_sample)
REM_ROWS = [
    # PoisonPy — 155 vuln, 155 ground truths
    ("PoisonPy",  "Semgrep --autofix", 155,  7,  7,  6,  6,  6,  5,  6,
     0.045, 1.000, 0.714, 0.032, 0.857, 0.039, 0.815, 4717.9),
    ("PoisonPy",  "PatchitPy (bash)",  155,  0,  0,  0,  0,  0,  0,  0,
     0.000, 0.000, 0.000, 0.000, 0.000, 0.000, None,  3968.5),
    ("PoisonPy",  "Redlyne v0.1.2",    155, 58, 58, 56, 54, 33, 52, 33,
     0.374, 1.000, 0.897, 0.335, 0.569, 0.213, 0.696, 2.7),

    # SafeCoder — 526 vuln, 526 ground truths
    ("SafeCoder", "Semgrep --autofix", 526, 15, 15,  1, 15,  6,  1,  6,
     0.029, 1.000, 0.067, 0.002, 0.400, 0.011, 0.413, 5019.2),
    ("SafeCoder", "PatchitPy (bash)",  526,  0,  0,  0,  0,  0,  0,  0,
     0.000, 0.000, 0.000, 0.000, 0.000, 0.000, None,  4381.5),
    ("SafeCoder", "Redlyne v0.1.2",    526, 101, 95, 72, 99, 42, 70, 42,
     0.192, 0.941, 0.693, 0.133, 0.416, 0.080, 0.492, 9.9),
]

REM_DATASET_META = {
    "PoisonPy":  ("Cotroneo et al., ICPC 2024",
                  "155 vulnerable samples, each paired with a hand-written `code_clean` ground truth. Synthetic but tight per-CWE coverage."),
    "SafeCoder": ("He et al., ICML 2024",
                  "526 real commit-based vulnerable Python functions, each paired with its `func_src_after` (the fix as it appeared in the public commit). Production-grade ground truth."),
}

# Metric rows for each dataset block on the Headline sheet.
# Tuples: (label, tuple-index, fmt, is-percentage)
REM_METRIC_ROWS = [
    ("Total samples",                  2,  "0",     False),
    ("Applied (count)",                3,  "0",     False),
    ("Applied rate",                  10, "0.0%",   True),
    ("Syntax-safe (count)",            4,  "0",     False),
    ("Syntax-safe / applied",         11, "0.0%",   True),
    ("Targeted-clean (count)",         5,  "0",     False),
    ("Regression-free (count)",        6,  "0",     False),
    ("Cross-clean (count)",            7,  "0",     False),
    ("Targeted-full (count)",          8,  "0",     False),
    ("Targeted-full / total",         13, "0.0%",   True),
    ("Targeted-full / applied",       12, "0.0%",   True),
    ("Strict-full (count)",            9,  "0",     False),
    ("Strict-full / total",           15, "0.0%",   True),
    ("Strict-full / applied",         14, "0.0%",   True),
    ("Mean similarity to GT",         16, "0.00",   False),
    ("Speed (ms / sample)",           17, "0.0",    False),
]


def build_remediation_headline(wb: Workbook):
    ws = wb.create_sheet("Remediation — Headline")
    ws["A1"] = "Remediation — every metric per dataset × tool"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REM_TOOLS) + 1)
    ws["A2"] = ("One block per dataset. Five core properties per patch: Applied → "
                "Syntax-safe → Targeted-clean (the rules with remediation that fired pre-patch "
                "stop firing) → Regression-free (no NEW rule IDs introduced) → Cross-clean "
                "(zero findings overall, strictest). Two composite 'fully clean' definitions: "
                "Targeted-full (headline) and Strict-full (whole file vuln-free post-patch).")
    ws["A2"].font = FONT_BODY
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(REM_TOOLS) + 1)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    lookup = {(r[0], r[1]): r for r in REM_ROWS}

    cur_row = 4
    for ds in REM_DATASETS:
        cite, descr = REM_DATASET_META[ds]

        # Dataset header
        title_cell = ws.cell(row=cur_row, column=1, value=f"{ds}  —  {cite}")
        title_cell.font = FONT_HEAD
        title_cell.fill = FILL_HEAD
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=len(REM_TOOLS) + 1)
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Dataset description
        desc_cell = ws.cell(row=cur_row, column=1, value=descr)
        desc_cell.font = FONT_BODY
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=len(REM_TOOLS) + 1)
        ws.row_dimensions[cur_row].height = 36
        cur_row += 1

        # Tool header row for this block
        ws.cell(row=cur_row, column=1, value="Metric")
        for c, tool in enumerate(REM_TOOLS, start=2):
            ws.cell(row=cur_row, column=c, value=tool)
        style_header_row(ws, cur_row, len(REM_TOOLS) + 1)
        cur_row += 1

        # Metric rows
        for metric_label, tup_idx, fmt, _ in REM_METRIC_ROWS:
            ws.cell(row=cur_row, column=1, value=metric_label)
            ws.cell(row=cur_row, column=1).font = FONT_BOLD
            ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=cur_row, column=1).border = BORDER
            for c, tool in enumerate(REM_TOOLS, start=2):
                r = lookup.get((ds, tool))
                v = r[tup_idx] if r else None
                cell = ws.cell(row=cur_row, column=c, value=v)
                fill = FILL_REDLY if tool.startswith("Redlyne") else None
                style_data_cell(cell, fill=fill, num_format=fmt)
            cur_row += 1

        cur_row += 1  # spacer

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 50
    ws.column_dimensions["A"].width = 28
    for c in range(2, len(REM_TOOLS) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 22
    ws.freeze_panes = "B4"


def build_remediation_full(wb: Workbook):
    ws = wb.create_sheet("Remediation — Full data")
    ws["A1"] = "Remediation — full long-format results"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)

    headers = [
        "Dataset", "Tool", "Total", "Applied", "Syntax-safe",
        "Targeted-clean", "Regression-free", "Cross-clean",
        "Targeted-full", "Strict-full",
        "Applied rate", "Syntax-safe / applied",
        "Targeted-full / applied", "Targeted-full / total",
        "Strict-full / applied",   "Strict-full / total",
        "Similarity to GT", "Speed (ms / sample)",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))

    # Column number formats (1-indexed): col 1-2 text, 3-10 counts, 11-16 %,
    # 17 decimal sim, 18 ms.
    pct_cols = {11, 12, 13, 14, 15, 16}
    count_cols = {3, 4, 5, 6, 7, 8, 9, 10}

    for i, r in enumerate(REM_ROWS, start=4):
        tool = r[1]
        fill = FILL_REDLY if tool.startswith("Redlyne") else None
        for c, v in enumerate(r, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            if c in (1, 2):
                cell.font = FONT_BOLD if c == 1 else FONT_BODY
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = BORDER
                if fill is not None:
                    cell.fill = fill
            elif c in count_cols:
                style_data_cell(cell, fill=fill, num_format="0")
            elif c in pct_cols:
                style_data_cell(cell, fill=fill, num_format="0.0%")
            elif c == 17:  # similarity
                style_data_cell(cell, fill=fill, num_format="0.00")
            elif c == 18:  # ms
                style_data_cell(cell, fill=fill, num_format="0.0")

    ws.auto_filter.ref = f"A3:R{3 + len(REM_ROWS)}"
    ws.freeze_panes = "C4"
    autosize(ws, max_w=24)


# ---------------------------------------------------------------------------
def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Drop the default sheet
    wb.remove(wb.active)
    build_headline(wb)
    build_full(wb)
    build_remediation_headline(wb)
    build_remediation_full(wb)
    build_methodology(wb)
    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
